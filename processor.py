import streamlit as st
import openpyxl
import pdfplumber
import re
from io import BytesIO

# 🚀 Dedicated parsers import
from parser_welspun import extract_welspun_items, map_items_to_excel_dynamic
from parser_bkt import extract_bkt_items
from parser_polycab import extract_polycab_items, map_polycab_items_to_excel_dynamic
from parser_vapi_welspun import extract_vapi_welspun_items, map_vapi_welspun_items_to_excel_dynamic

from shipper_data import fetch_data_from_google_sheet, ensure_default_shipper
from pdf_engine import apply_rule_filter, extract_header_value
from google_sheet_sync import load_template_from_sheet
from supporting_engine import extract_data_from_supporting_file

def render_processor():
    fetch_data_from_google_sheet()
    ensure_default_shipper()
    
    st.header("📤 Invoice Processing Zone (Multi-Document)")
    st.caption("इनवॉइस PDF या Excel के साथ-साथ GST Invoice और DEEC Declaration अपलोड करने और पर्टिकुलर सेल/कॉलम में भेजने का ज़ोन.")
    
    # 🚀 1. शिपर लिस्ट को अल्फाबेटिकल (A से Z) क्रम में सॉर्ट किया गया है
    shippers_list = sorted(list(st.session_state["shipper_database"].keys()))
    
    if shippers_list:
        # 🚀 2. index=None और placeholder से यह डिफ़ॉल्ट रूप से blank रहेगा और टाइप करके भी खोजा जा सकेगा
        selected_shipper = st.selectbox(
            "किस शिपर का इनवॉइस प्रोसेस करना है?", 
            shippers_list, 
            index=None, 
            placeholder="शिपर का नाम टाइप करें या चुनें..."
        )
        
        if selected_shipper:
            shipper_info = st.session_state["shipper_database"][selected_shipper]
            
            if f"inv_count_{selected_shipper}" not in st.session_state:
                st.session_state[f"inv_count_{selected_shipper}"] = 1
            
            inv_count = st.session_state[f"inv_count_{selected_shipper}"]
            
            st.subheader("📑 Upload Invoices & Supporting Documents")
            
            uploaded_batches = []
            for i in range(inv_count):
                st.markdown(f"#### ➡️ Invoice Set #{i+1}")
                col_inv, col_gst, col_deec = st.columns(3)
                
                with col_inv:
                    pdf_f = st.file_uploader(f" मुख्य इनवॉइस (PDF / Excel) #{i+1}", type=["pdf", "xlsx", "xls"], key=f"inv_pdf_{selected_shipper}_{i}")
                with col_gst:
                    gst_f = st.file_uploader(f" GST Invoice #{i+1} (PDF/Excel)", type=["pdf", "xlsx", "xls"], key=f"gst_file_{selected_shipper}_{i}")
                with col_deec:
                    deec_f = st.file_uploader(f" DEEC Decl. #{i+1} (PDF/Excel)", type=["pdf", "xlsx", "xls"], key=f"deec_file_{selected_shipper}_{i}")
                    
                uploaded_batches.append((i+1, pdf_f, gst_f, deec_f))
                st.write("---")
            
            col_b1, col_b2, col_space = st.columns([2, 2, 6])
            with col_b1:
                if inv_count < 10:
                    if st.button("➕ Add Invoice Set", key=f"add_btn_{selected_shipper}", use_container_width=True):
                        st.session_state[f"inv_count_{selected_shipper}"] += 1
                        st.rerun()
            with col_b2:
                if inv_count > 1:
                    if st.button("➖ Remove Last", key=f"rem_btn_{selected_shipper}", use_container_width=True):
                        st.session_state[f"inv_count_{selected_shipper}"] -= 1
                        st.rerun()
                        
            st.write("---")
            
            valid_batches = [b for b in uploaded_batches if b[1] is not None]
            
            if valid_batches and st.button("🚀 Process & Generate Excel with Supporting Docs", type="primary", use_container_width=True):
                with st.spinner(f"कुल {len(valid_batches)} इनवॉइस सेट प्रोसेस हो रहे हैं..."):
                    rules = shipper_info.get("mapping_rules", {})
                    item_table_rules = shipper_info.get("item_table_rules", {})
                    
                    # 🚀 शिपर के हिसाब से असाइंड पार्सर रूल पढ़ें
                    assigned_parser = shipper_info.get("item_table_rule_name", "parser_welspun").strip().lower()
                    
                    igst_cfg = shipper_info.get("igst_config", {})
                    lut_kws = igst_cfg.get("lut_keywords", "")
                    paid_kws = igst_cfg.get("paid_keywords", "")
                    
                    wb = load_template_from_sheet(selected_shipper)
                    if wb is None:
                        wb = openpyxl.Workbook()
                        
                    ws = wb["INV"] if "INV" in wb.sheetnames else wb.active
                    
                    first_inv_no = "INV"
                    overall_item_sr = 1
                    excel_write_row = 2
                    
                    for inv_sr_number, inv_file, gst_file, deec_file in valid_batches:
                        pdf_text = ""
                        pdf_lines = []
                        
                        if inv_file:
                            file_bytes_cache = inv_file.getvalue()
                            st.session_state["cached_pdf_bytes"] = file_bytes_cache
                        
                        file_name_lower = inv_file.name.lower()
                        if file_name_lower.endswith(".pdf"):
                            with pdfplumber.open(BytesIO(st.session_state["cached_pdf_bytes"])) as pdf:
                                for page in pdf.pages:
                                    t = page.extract_text()
                                    if t:
                                        pdf_text += t + "\n"
                                        pdf_lines.extend(t.split("\n"))
                        else:
                            excel_text, _ = extract_data_from_supporting_file(inv_file)
                            if excel_text:
                                pdf_text = excel_text
                                pdf_lines = excel_text.split("\n")
                        
                        gst_text, _ = extract_data_from_supporting_file(gst_file) if gst_file else ("", None)
                        deec_text, _ = extract_data_from_supporting_file(deec_file) if deec_file else ("", None)
                        
                        current_inv_number = f"INV_{inv_sr_number}"
                        current_inv_date = ""
                        inv_data_dict = {}
                        
                        summary_row = 1 + inv_sr_number
                        
                        for field, r_info in rules.items():
                            kw = r_info.get("keyword", "").strip()
                            if kw.startswith("'") and len(kw) > 1:
                                kw = kw[1:].strip()
                                
                            pos = r_info.get("position", "Right (आगे)")
                            target_cell = r_info.get("cell", "").strip().upper()
                            mode = r_info.get("match_mode", "Exact Word")
                            stop_kw = r_info.get("stop_kw", "").strip()
                            flt = r_info.get("filter", "None")
                            fallback_val = r_info.get("fallback", "").strip()
                            doc_source = r_info.get("logic", "Main Invoice")
                            
                            # 🚀 चेक करें कि क्या इस फील्ड के लिए कोई कस्टम 'extracted_logic' (Python/Regex) सेव है
                            extracted_logic = r_info.get("extracted_logic", "").strip()
                            found_val = None
                            
                            target_lines, target_full_text = pdf_lines, pdf_text
                            if "gst" in doc_source.lower() and gst_file:
                                target_lines = gst_text.split("\n")
                                target_full_text = gst_text
                            elif "deec" in doc_source.lower() and deec_file:
                                target_lines = deec_text.split("\n")
                                target_full_text = deec_text
                                
                            # यदि कस्टम लॉजिक उपलब्ध है, तो उसे सीधे रन करें
                            if extracted_logic:
                                try:
                                    clean_code = extracted_logic.replace("```python", "").replace("```", "").strip()
                                    local_vars = {"text": target_full_text, "lines": target_lines, "re": re}
                                    exec(clean_code, {}, local_vars)
                                    found_val = local_vars.get("value", None)
                                except Exception as e:
                                    found_val = None
                                    
                            # यदि कस्टम लॉजिक से वैल्यू न मिले, तो डिफ़ॉल्ट एक्सट्रैक्शन मेथड इस्तेमाल करें
                            if not found_val or not str(found_val).strip():
                                pdf_bytes = st.session_state.get("cached_pdf_bytes", None)
                                found_val = extract_header_value(target_lines, target_full_text, kw, pos, mode, stop_kw, flt, field_label=field, pdf_bytes=pdf_bytes)
                            
                            if not found_val or not str(found_val).strip():
                                if fallback_val:
                                    found_val = fallback_val
                                    
                            inv_data_dict[field.lower()] = found_val
                            
                            if target_cell and "dynamic" not in target_cell.lower():
                                try:
                                    if "\n" in str(found_val):
                                        col_letters = re.findall(r'[A-Za-z]+', target_cell)[0].upper()
                                        start_row_num = int(re.findall(r'\d+', target_cell)[0]) if re.findall(r'\d+', target_cell) else summary_row
                                        
                                        lines = str(found_val).split("\n")
                                        for idx, line_val in enumerate(lines):
                                            current_row = start_row_num + idx
                                            ws[f"{col_letters}{current_row}"] = line_val.strip()
                                    else:
                                        if target_cell.isalpha():
                                            cell_to_write = f"{target_cell}{summary_row}"
                                        else:
                                            cell_to_write = target_cell
                                        ws[cell_to_write] = found_val
                                except Exception:
                                    pass
                            
                            if "inv. no" in field.lower() or "invoice no" in field.lower():
                                if found_val:
                                    current_inv_number = found_val
                                    if inv_sr_number == 1: first_inv_no = found_val
                            
                            if "date" in field.lower() or "dt" in field.lower():
                                d_match = re.search(r'\b\d{2}[./-]\d{2}[./-]\d{4}\b', str(found_val))
                                if d_match:
                                    current_inv_date = d_match.group(0).replace(".", "/").replace("-", "/")
                                elif found_val and not str(found_val).lower().startswith("inv"):
                                    current_inv_date = found_val

                        ws[f"AH{summary_row}"] = inv_sr_number
                        ws[f"AI{summary_row}"] = current_inv_number
                        
                        if current_inv_date:
                            ws[f"AJ{summary_row}"] = current_inv_date

                        resolved_item_rules = {}
                        for i_name, i_info in item_table_rules.items():
                            i_type = i_info.get("type", "")
                            i_rule = i_info.get("rule", "")
                            if i_rule.startswith("'") and len(i_rule) > 1:
                                i_rule = i_rule[1:].strip()
                            i_col = i_info.get("col", "K")
                            
                            actual_rule_val = i_rule
                            if i_type == "Header Field Mapping":
                                matched_header_key = i_rule.lower()
                                if matched_header_key in inv_data_dict:
                                    actual_rule_val = inv_data_dict[matched_header_key]
                            
                            resolved_item_rules[i_name] = {
                                "col": i_col,
                                "type": i_type if i_type != "Header Field Mapping" else "Constant Text",
                                "rule": actual_rule_val
                            }

                        # 🚀 सटीक पार्सर कॉलिंग (Vapi Welspun, Polycab, BKT या Welspun)[cite: 4]
                        if assigned_parser == "parser_bkt":
                            parsed_items = extract_bkt_items(pdf_lines)
                        elif assigned_parser == "parser_polycab":
                            parsed_items = extract_polycab_items(pdf_lines, pdf_text=pdf_text)
                        elif assigned_parser == "parser_vapi_welspun":
                            parsed_items = extract_vapi_welspun_items(pdf_lines, pdf_text=pdf_text)
                        elif assigned_parser == "parser_welspun":
                            parsed_items = extract_welspun_items(pdf_lines, pdf_text=pdf_text)
                        else:
                            parsed_items = extract_welspun_items(pdf_lines, pdf_text=pdf_text)
                        
                        # 🚀 मैपिंग फंक्शन कॉलिंग
                        if assigned_parser == "parser_polycab":
                            ws, overall_item_sr, excel_write_row = map_polycab_items_to_excel_dynamic(
                                ws, parsed_items, resolved_item_rules,
                                inv_sr_no=inv_sr_number, 
                                start_overall_sr=overall_item_sr, 
                                start_excel_row=excel_write_row, 
                                default_invoice_no=current_inv_number, 
                                default_invoice_date=current_inv_date,
                                pdf_text=pdf_text,
                                lut_kws=lut_kws,
                                paid_kws=paid_kws,
                                parser_rule=assigned_parser
                            )
                        elif assigned_parser == "parser_vapi_welspun":
                            ws, overall_item_sr, excel_write_row = map_vapi_welspun_items_to_excel_dynamic(
                                ws, parsed_items, resolved_item_rules,
                                inv_sr_no=inv_sr_number, 
                                start_overall_sr=overall_item_sr, 
                                start_excel_row=excel_write_row, 
                                default_invoice_no=current_inv_number, 
                                default_invoice_date=current_inv_date,
                                pdf_text=pdf_text,
                                lut_kws=lut_kws,
                                paid_kws=paid_kws,
                                parser_rule=assigned_parser
                            )
                        else:
                            ws, overall_item_sr, excel_write_row = map_items_to_excel_dynamic(
                                ws, parsed_items, resolved_item_rules,
                                inv_sr_no=inv_sr_number, 
                                start_overall_sr=overall_item_sr, 
                                start_excel_row=excel_write_row, 
                                default_invoice_no=current_inv_number, 
                                default_invoice_date=current_inv_date,
                                pdf_text=pdf_text,
                                lut_kws=lut_kws,
                                paid_kws=paid_kws,
                                parser_rule=assigned_parser
                            )

                    output = BytesIO()
                    wb.save(output)
                    
                    short_shipper = selected_shipper.split(" ")[0].lower()
                    clean_inv = re.sub(r'[\\/*?:"<>|]', "", first_inv_no)
                    final_filename = f"{clean_inv}_{short_shipper}_MultiDoc.xlsx"
                    
                    st.session_state["processed_file_ready"] = {"filename": final_filename, "data": output.getvalue()}
                    st.success(f"🎉 सफलता! सपोर्टिंग डॉक्यूमेंट्स के साथ फाइल '{final_filename}' तैयार है!")
            
            if st.session_state.get("processed_file_ready", None):
                st.download_button(
                    label=f"📥 {st.session_state['processed_file_ready']['filename']} डाउनलोड करें",
                    data=st.session_state['processed_file_ready']['data'],
                    file_name=st.session_state['processed_file_ready']['filename'],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                st.session_state["processed_file_ready"] = None
