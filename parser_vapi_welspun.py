import re
import streamlit as st
import pdfplumber
from io import BytesIO
from openpyxl.styles import Alignment
from pdf_engine import apply_value_replacement, extract_header_value

def extract_all_commodities_from_text(pdf_text):
    commodities = []
    pattern = r'(\d{8})\s*[:\-]\s*(.+)'
    matches = re.findall(pattern, pdf_text)
    for hsn, desc in matches:
        full_desc = f"{hsn}: {desc.strip()}"
        commodities.append(full_desc)
    return commodities

def extract_vapi_welspun_items(pdf_lines, pdf_text=""):
    parsed_items = []
    box_commodities = extract_all_commodities_from_text(pdf_text)
    
    cached_bytes = st.session_state.get("cached_pdf_bytes", None)
    if cached_bytes:
        try:
            with pdfplumber.open(BytesIO(cached_bytes)) as pdf:
                for page in pdf.pages:
                    tables = page.extract_tables()
                    for table in tables:
                        for row in table:
                            if row and len(row) >= 5:
                                clean_cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip() != ""]
                                if not clean_cells:
                                    continue
                                    
                                hs_code = ""
                                hs_index = -1
                                for idx, cell in enumerate(clean_cells):
                                    if re.fullmatch(r'\d{8}', cell.replace(",", "").strip()):
                                        hs_code = cell.replace(",", "").strip()
                                        hs_index = idx
                                        break
                                        
                                if not hs_code:
                                    continue
                                    
                                dbk_sr = clean_cells[hs_index - 1] if hs_index > 0 else ""

                                # डायनेमिक डिस्क्रिप्शन
                                description_text = ""
                                if hs_index != -1 and len(clean_cells) > hs_index + 1:
                                    desc_parts = []
                                    for idx in range(hs_index + 1, len(clean_cells)):
                                        cell_val = clean_cells[idx]
                                        if re.fullmatch(r'\d+\.\d+', cell_val) or re.fullmatch(r'\d+\s*[xX]\s*\d+', cell_val):
                                            break
                                        if cell_val.isdigit() and int(cell_val) > 99 and len(desc_parts) > 0:
                                            break
                                        desc_parts.append(cell_val)
                                    description_text = " ".join(desc_parts).strip()

                                if not description_text and hs_index > 0:
                                    desc_candidates = [clean_cells[i] for i in range(0, hs_index) if not clean_cells[i].isdigit()]
                                    description_text = " ".join(desc_candidates).strip()

                                # 🚀 फिक्स: SQMTR और Net Wt को उनके क्रम (Order) से सेट करना ताकि कभी इंटरचेंज न हों
                                net_wt, qty, rate, amount_usd, taxable_inr, igst_per, igst_amt, sqmtr = "", "", "", "", "", "", "", ""
                                
                                if len(clean_cells) >= 4:
                                    igst_amt = clean_cells[-1]
                                    igst_per = clean_cells[-2]
                                    taxable_inr = clean_cells[-3]
                                    amount_usd = clean_cells[-4]

                                decimal_3_values = []
                                for cell in clean_cells:
                                    clean_c = cell.replace(",", "").strip()
                                    if re.fullmatch(r'\d+\.\d{3}', clean_c):
                                        decimal_3_values.append(cell)
                                    elif re.fullmatch(r'\d+\.\d{5}', clean_c):
                                        if not rate:
                                            rate = cell
                                    elif clean_c.isdigit() and int(clean_c) > 0 and cell != hs_code and cell != dbk_sr:
                                        if not qty:
                                            qty = cell

                                # यदि एक से ज्यादा 3-डेसिमल वैल्यू हैं (जैसे SQMTR और Net Wt दोनों)
                                if len(decimal_3_values) >= 2:
                                    sqmtr = decimal_3_values[0]   # पहली वैल्यू SQMTR बनेगी
                                    net_wt = decimal_3_values[1]  # दूसरी वैल्यू हमेशा Net Wt बनेगी
                                elif len(decimal_3_values) == 1:
                                    net_wt = decimal_3_values[0]  # अगर सिर्फ एक है तो वह Net Wt होगी

                                item_dict = {f"col_{i}": (str(row[i]).strip() if i < len(row) and row[i] else "") for i in range(len(row))}
                                item_dict.update({
                                    "dbk_sr": dbk_sr,
                                    "hs_code": hs_code,
                                    "description_text": description_text,
                                    "net_wt": net_wt,
                                    "qty": qty,
                                    "rate": rate,
                                    "amount_usd": amount_usd,
                                    "amount_inr": taxable_inr,
                                    "igst_per": igst_per if igst_per else "5.00",
                                    "igst_amt": igst_amt,
                                    "sqmtr": sqmtr,
                                    "box_commodities": box_commodities
                                })
                                parsed_items.append(item_dict)
        except Exception as e:
            st.error(f"Pattern Parser Error: {str(e)}")

    if not parsed_items:
        parsed_items.append({"dbk_sr": "", "hs_code": "", "description_text": "", "net_wt": "", "qty": "", "rate": "", "amount_usd": "", "amount_inr": "", "igst_per": "5.00", "igst_amt": "", "sqmtr": "", "box_commodities": []})

    return parsed_items


def map_vapi_welspun_items_to_excel_dynamic(ws, parsed_items, item_rules, inv_sr_no=1, start_overall_sr=1, start_excel_row=2, default_invoice_no="", default_invoice_date="", pdf_text="", lut_kws="", paid_kws="", parser_rule=""):
    curr_row = start_excel_row
    overall_sr = start_overall_sr
    
    pdf_text_upper = str(pdf_text).upper()
    pdf_lines = str(pdf_text).split("\n")
    
    l_keywords = [k.strip().upper() for k in str(lut_kws).split(",") if k.strip()]
    p_keywords = [k.strip().upper() for k in str(paid_kws).split(",") if k.strip()]
    
    matched_lut = any(kw.replace("NO.", "").replace(".", "").strip() in pdf_text_upper for kw in l_keywords if kw.strip())
    matched_paid = any(kw.replace(".", "").strip() in pdf_text_upper for kw in p_keywords if kw.strip())

    v_column_value = "LUT" if matched_lut else ("P" if matched_paid else "LUT")

    max_rows = len(parsed_items)

    first_item = parsed_items[0] if parsed_items else {}
    all_comms = first_item.get("box_commodities", [])

    for item_idx in range(max_rows):
        item_sr_no = item_idx + 1
        item = parsed_items[item_idx] if item_idx < len(parsed_items) else {}
        
        ws[f"G{curr_row}"] = inv_sr_no                    
        ws[f"H{curr_row}"] = item_sr_no                                      
        ws[f"V{curr_row}"] = v_column_value               
        
        ws[f"I{curr_row}"] = default_invoice_no
        if default_invoice_date and not "ROSC" in str(default_invoice_date):
            ws[f"J{curr_row}"] = default_invoice_date

        if all_comms:
            cell_ref_bs = f"BS{curr_row}"
            if item_idx < len(all_comms):
                ws[cell_ref_bs] = all_comms[item_idx]
            else:
                ws[cell_ref_bs] = ""
            ws[cell_ref_bs].alignment = Alignment(wrap_text=True)

        for field_name, r_info in item_rules.items():
            col_letter = r_info.get("col", "").strip().upper()
            rule_type_raw = str(r_info.get("type", "PDF Row Item")).strip()
            rule_val = str(r_info.get("rule", "")).strip()
            
            if not col_letter or col_letter in ["V", "BR", "BS", "S", "J"]:
                continue
                
            if "extract" in rule_type_raw.lower() or "box" in rule_type_raw.lower() or "header" in rule_type_raw.lower() or col_letter in ["BW", "BY"]:
                cached_bytes = st.session_state.get("cached_pdf_bytes", None)
                extracted_val = extract_header_value(pdf_lines, pdf_text, rule_val, "📦 Extract Inside Box (डब्बे के अंदर का टेक्स्ट)", "Exact Word", "", "None", field_label=field_name, pdf_bytes=cached_bytes)
                if not extracted_val or not extracted_val.strip():
                    extracted_val = extract_header_value(pdf_lines, pdf_text, rule_val, "Right (आगे)", "Exact Word", "", "None", field_label=field_name)
                
                if extracted_val and "\n" in str(extracted_val):
                    lines = [l.strip() for l in str(extracted_val).split("\n") if l.strip()]
                    ws[f"{col_letter}{curr_row}"] = lines[item_idx] if item_idx < len(lines) else ""
                else:
                    ws[f"{col_letter}{curr_row}"] = extracted_val if item_idx == 0 else ""

        for field_name, r_info in item_rules.items():
            col_letter = r_info.get("col", "").strip().upper()
            rule_type_raw = str(r_info.get("type", "PDF Row Item")).strip()
            rule_val = str(r_info.get("rule", "")).strip()
            rule_val_lower = rule_val.lower()
            
            if not col_letter or col_letter in ["V", "I", "J", "G", "BR", "BS", "M"]:
                continue
            
            if "extract" in rule_type_raw.lower() or "box" in rule_type_raw.lower() or "header" in rule_type_raw.lower():
                continue
                
            cell_ref = f"{col_letter}{curr_row}"
            raw_val = ""
            clean_rule_val = rule_val_lower.replace("col_", "").strip()
            
            if clean_rule_val.isdigit() and col_letter in ["K", "S", "BU"]:
                col_idx = int(clean_rule_val)
                raw_val = item.get(f"col_{col_idx}", "")
                if col_letter == "S":
                    raw_val = f"{raw_val}B" if raw_val and not str(raw_val).upper().endswith("B") else raw_val

            elif col_letter == "K" or "hs" in rule_val_lower or "ritc" in rule_val_lower:
                raw_val = item.get("hs_code", "")
            elif col_letter == "BU" or "desc" in rule_val_lower:
                raw_val = item.get("description_text", "")
            elif col_letter == "S" or "dbk" in rule_val_lower:
                dbk = item.get("dbk_sr", "")
                raw_val = f"{dbk}B" if dbk and not dbk.upper().endswith("B") else dbk
            elif col_letter == "AB" or "wt" in rule_val_lower or "weight" in rule_val_lower:
                raw_val = item.get("net_wt", "")
            elif col_letter == "N" or "qty" in rule_val_lower or "quantity" in rule_val_lower:
                raw_val = item.get("qty", "")
            elif col_letter == "P" or "rate" in rule_val_lower:
                raw_val = item.get("rate", "")
            elif col_letter == "Q" or ("amount" in rule_val_lower and "usd" in rule_val_lower):
                raw_val = item.get("amount_usd", "")
            elif col_letter == "W" or "taxable" in rule_val_lower or "inr" in rule_val_lower:
                raw_val = item.get("amount_inr", "")
            elif col_letter == "X" or "igst%" in rule_val_lower or "igst per" in rule_val_lower:
                raw_val = item.get("igst_per", "5.00")
            elif col_letter == "Y" or "igst amount" in rule_val_lower:
                raw_val = item.get("igst_amt", "")
            elif col_letter == "Z" or "sqmtr" in rule_val_lower:
                raw_val = item.get("sqmtr", "")
            else:
                if clean_rule_val.isdigit():
                    col_idx = int(clean_rule_val)
                    raw_val = item.get(f"col_{col_idx}", "")

            if "=" in rule_val:
                raw_val = apply_value_replacement(str(raw_val), rule_val)

            try:
                if col_letter in ["S", "K", "BU"]:
                    ws[cell_ref] = str(raw_val).replace("\n", " ")
                else:
                    clean_num = str(raw_val).replace(",", "").replace("\n", "").strip()
                    ws[cell_ref] = float(clean_num) if clean_num else 0.0
            except:
                ws[cell_ref] = raw_val
                    
        curr_row += 1
        overall_sr += 1
        
    return ws, overall_sr, curr_row
